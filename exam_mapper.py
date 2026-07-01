"""
Exam Question Mapper — Gold Version
=====================================
Parses multiple-choice exam PDFs (Forms A/B/C/D...), maps question numbers
and choice orders across forms, and detects the correct answer.

All fixes included:
  - Global one-to-one matching (no question missed due to match contention)
  - OMIT="Yes" for red-box not-graded questions
  - Multi-answer detection (two bolded choices captured as "A,C")
  - Per-form continuity check (no number jumps like 5→7)
  - Completeness guard (no question silently dropped)
  - Same-stem different-choices split into separate variant rows
  - Dynamic form-letter assignment (any number of early/late forms)
  - L1/L10 column headers correct
"""

import os, re, sys, difflib, warnings, logging
import pypdf
warnings.filterwarnings("ignore", category=pypdf.errors.PdfReadWarning)
logging.getLogger("pypdf").setLevel(logging.ERROR)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── LOGGING ───────────────────────────────────────────────────────────────────
_logger = logging.getLogger("exam_mapper")
if not _logger.handlers:
    _h = logging.StreamHandler(sys.stdout)
    _h.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)-7s %(message)s",
                                      datefmt="%H:%M:%S"))
    _logger.addHandler(_h)
_logger.setLevel(os.environ.get("EXAM_MAPPER_LOGLEVEL", "WARNING").upper())
_logger.propagate = False

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
ANSWER_KEY_FORM_A: dict = {}
FORMS = ["A","B","C","D","E","F","G","H","I","J","K","L"]
TICK_CHARS    = {"\u2714","\u2713","\u2705","\u2611","✔","✓","☑","✅",
                 "\x14","\x13","\x12","\u2714\ufe0f"}
DINGBAT_FONTS = {"DINGBAT","WINGDING","ZAPF","SYMBOL"}
BOLD_KEYWORDS = ["BOLD","BLACK","HEAVY","DEMI"]

MATCH_THRESHOLD        = 0.65
WORD_OVERLAP_THRESHOLD = 0.40
VARIANT_THRESHOLD      = 0.999
CHOICE_OVERLAP_MIN     = 0.40
MAX_QUESTION_NUMBER    = 150

# Side channel: omit notes detected but not confidently attachable to a specific
# question (e.g. a red box that floated in reading order). Keyed by filename;
# surfaced in the issues report for manual review so a real "not graded"
# instruction is never silently dropped, and never turned into a false omit.
_UNATTACHED_OMITS = {}

OMIT_PHRASES = re.compile(
    r'not\s+graded|'
    r'(?:not|n[\'o]t)\s+count(?:ing|ed)?\s+(?:this\s+)?(?:question|problem|one|it)|'
    r'(?:we\s+are\s+)?not\s+count(?:ing)?\s+this|'
    r"(?:do(?:es)?\s*n[\'o]?t|won[\'o]?t|will\s+not)\s+count|"
    r'omit(?:ted)?\s+(?:this\s+)?(?:question|problem)|'
    r'(?:disregard|throw[n]?\s+out|thrown\s+out)\s+(?:this\s+)?(?:question|problem)|'
    r'no\s+credit|credit\s+(?:to|for)\s+all|credit\s+to\s+everyone|'
    r'giv(?:e|ing|en)\s+(?:everyone|all\s+students|everybody|full\s+credit)|'
    r'(?:given|gave|giving|receive[ds]?|got|get)\s+(?:full\s+)?credit|'
    r'everyone\s+(?:gets|got|was|were|receive[ds]?|gave|given|will)|'
    r'all\s+(?:students\s+)?(?:get|got|receive[ds]?|given)\s+(?:full\s+)?credit|'
    r'full\s+credit\s+(?:to|for)\s+all|'
    r'count(?:ed|s)?\s+as\s+correct\s+for\s+all',
    re.I
)

EXTRA_COLS = [
    "Version","Exam Seq Qno","Correct_Answer","Master Question Number.1",
    "VOICE","Vcode","TENSE","Tcode","FORM","Fcode",
    "PERSON","Pcode","CONSTR","Ccode",
    "L1","L2","L3","L4","L5","L6","L7","L8","L9","L10"
]

Q_SPLIT_RE = re.compile(
    r'(?:_{5,}[ \t]*|\xa0|(?:^|\n)[ \t]*|[ \t]{3,}'
    r'|(?<=[a-zA-Z%)])[ \t]{2}(?=[1-9]\d?[.:][ \t]))'
    r'([1-9]\d?)[.:][ \t]+'
)

# ── RED-BOX DETECTION ─────────────────────────────────────────────────────────

def _detect_red_boxes(pdf_path: str) -> dict:
    """Scan PDF content streams for red filled rectangles.

    Returns {q_num_str: 'omit' | 'two_answers'}.
    'omit'        -> professor marked question as not graded  -> OMIT="Yes"
    'two_answers' -> two correct answers exist                -> capture both
    """
    try:
        reader = pypdf.PdfReader(pdf_path)
    except Exception:
        return {}
    results = {}
    for page in reader.pages:
        content_obj = page.get('/Contents')
        if not content_obj:
            continue
        try:
            if isinstance(content_obj, pypdf.generic.ArrayObject):
                raw = b''.join(o.get_object().get_data() for o in content_obj)
            else:
                raw = content_obj.get_object().get_data()
        except Exception:
            continue
        stream   = raw.decode('latin-1', errors='replace')
        page_txt = page.extract_text() or ""

        # Find reddish fill commands: r g b rg where r>0.55, g<0.55, b<0.55
        for mc in re.finditer(r'([\d.]+)[ \t]+([\d.]+)[ \t]+([\d.]+)[ \t]+(?:RG|rg)', stream):
            try:
                r, g, b = float(mc.group(1)), float(mc.group(2)), float(mc.group(3))
            except ValueError:
                continue
            if not (r > 0.55 and g < 0.55 and b < 0.55):
                continue
            rpos   = mc.start()
            window = stream[rpos: rpos + 800]
            if not re.search(r're\b.*?\bf\b', window, re.S):
                continue
            before = stream[:rpos]
            q_num  = None
            for mq in re.finditer(r'(?<![.\d])(\d{1,2})\.', before):
                n = int(mq.group(1))
                if 1 <= n <= MAX_QUESTION_NUMBER:
                    q_num = str(n)
            if not q_num or q_num in results:
                continue
            local = stream[max(0, rpos-300): rpos+300]
            if OMIT_PHRASES.search(page_txt) or OMIT_PHRASES.search(local):
                results[q_num] = 'omit'
            else:
                results[q_num] = 'two_answers'
    return results


# ── TEXT + BOLD EXTRACTION ────────────────────────────────────────────────────

def _is_bold(fd) -> bool:
    return bool(fd) and any(kw in str(fd.get("/BaseFont","")).upper() for kw in BOLD_KEYWORDS)

def _is_dingbat(fd) -> bool:
    return bool(fd) and any(kw in str(fd.get("/BaseFont","")).upper() for kw in DINGBAT_FONTS)


def extract_page_data(pdf_path: str) -> list:
    reader = pypdf.PdfReader(pdf_path)
    if len(reader.pages) > 1:
        p1 = reader.pages[0].extract_text() or ""
        if re.search(r'\bCHEM\s+\d{4}\b', p1, re.I):
            start = 1
        else:
            start = 0 if re.search(
                r'(?:^|\n|(?<=\s{2}))[ \t]{0,8}[1-9]\d?\.[ \t]{1,4}[A-Za-z]',
                p1, re.MULTILINE) else 1
    else:
        start = 0
    result = []
    for page in reader.pages[start:]:
        segs = []
        def vis(text, cm, tm, fd, fs):
            if text:
                if _is_dingbat(fd) and '4' in text:
                    text = text.replace('4', '\u2714')
                segs.append((text, _is_bold(fd)))
        try:
            page.extract_text(visitor_text=vis)
        except Exception:
            result.append((page.extract_text() or "", set())); continue
        plain = ""; bold = set()
        for txt, ib in segs:
            s = len(plain); plain += txt
            if ib: bold.update(range(s, s+len(txt)))
        result.append((plain, bold))
    return result


def extract_text(pdf_path: str) -> str:
    r = pypdf.PdfReader(pdf_path)
    return "\n".join(p.extract_text() or "" for p in r.pages)


# ── METADATA ──────────────────────────────────────────────────────────────────

def _date_to_semester(month: int, year: str) -> str:
    m = int(month)
    return f"{'Spring' if m<=5 else 'Summer' if m<=8 else 'Fall'} {year}"


def extract_metadata(text: str, filename: str) -> dict:
    meta = {"semester":"","course":"","exam":"","form":""}
    fn   = os.path.splitext(os.path.basename(filename))[0]
    m = re.search(r'(?:Exam\d*|Form|_v)([A-D])(?:_|\b)', fn, re.I)
    meta["form"] = m.group(1).upper() if m else ""
    if not meta["form"]:
        m = re.search(r'(?:EXAM\s+TYPE|TEST\s+FORM)\s*:\s*([A-D])', text[:600], re.I)
        if m: meta["form"] = m.group(1)
    m = re.search(r'\b(Fall|Spring|Summer|Winter)\s+(\d{2,4})\b', text[:800], re.I)
    if m:
        meta["semester"] = f"{m.group(1).capitalize()} {m.group(2)}"
    else:
        m = re.search(
            r'\b(January|February|March|April|May|June|July|August|'
            r'September|October|November|December|Jan|Feb|Mar|Apr|'
            r'Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},?\s+(\d{4})\b',
            text[:800], re.I)
        if m:
            MN = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
                  'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
                  'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,
                  'sep':9,'oct':10,'nov':11,'dec':12}
            mn = MN.get(m.group(1).lower(), 0)
            if mn: meta["semester"] = _date_to_semester(mn, m.group(2))
    m = re.search(r'\b([A-Z]{2,6}\s*\d{3,5}[A-Z]?)\b', text[:400])
    if m: meta["course"] = m.group(1).strip()
    m = re.search(r'\b((?:Exam|Midterm|Final|Quiz)\s*#?\d*)\b', text[:400], re.I)
    if m: meta["exam"] = m.group(1).strip()
    if not meta["semester"]:
        cm = re.search(r'\b(F|S|Su)(\d{2})\b', fn, re.I)
        if cm:
            pfx = cm.group(1).upper(); yr = int(cm.group(2))
            year = f"20{yr:02d}" if yr <= 30 else f"19{yr:02d}"
            s = {"F":"Fall","S":"Spring","SU":"Summer"}.get(pfx,"")
            if s: meta["semester"] = f"{s} {year}"
    return meta


# ── QUESTION PARSING ──────────────────────────────────────────────────────────

def _has_tick(s):  return any(c in s for c in TICK_CHARS)
def _strip_tick(s):
    for c in TICK_CHARS: s = s.replace(c,"")
    return s.strip()
def _norm(s): return re.sub(r'[^a-z0-9]','',s.lower())
def _tail(s):
    m = re.search(r'(?:what|calculate|determine|find|which|how|identify)\b.*$', s, re.I)
    if m: return re.sub(r'[^a-z0-9]','',m.group(0).lower())
    parts = re.split(r'[.?!]\s+', s.strip())
    t = parts[-1].strip() if parts else s
    if len(t) < 40 and len(s) > 40: t = s[-80:]
    return re.sub(r'[^a-z0-9]','',t.lower())
def _sim(a,b): return difflib.SequenceMatcher(None,a,b).ratio()
def _word_overlap(a,b):
    wa = set(w for w in re.findall(r'[a-z]{4,}', a.lower()))
    wb = set(w for w in re.findall(r'[a-z]{4,}', b.lower()))
    if not wa or not wb: return 0.0
    return len(wa&wb)/len(wa|wb)


def _detect_bold_answers(choice_offsets: dict, bold_idx: set) -> set:
    """Return ALL bold-marked choice letters (multi-answer aware)."""
    letters = sorted(choice_offsets.keys())
    if not letters: return set()
    mb = {}; tb = {}
    for k, L in enumerate(letters):
        s = choice_offsets[L]
        e = choice_offsets[letters[k+1]] if k+1<len(letters) else s+80
        me = min(s+3, e)
        mb[L] = sum(1 for i in range(s,me) if i in bold_idx)
        tb[L] = sum(1 for i in range(s,e)  if i in bold_idx)
    scores = {L: mb[L]*3+tb[L] for L in letters}
    if not any(scores.values()): return set()
    accepted = set()
    for L in letters:
        others = max((scores[l] for l in letters if l!=L), default=0)
        if mb[L] >= 2 or (tb[L] >= 2 and scores[L] > others):
            accepted.add(L)
    return accepted


def _mathbold_to_ascii(s):
    out = []
    for c in s:
        o = ord(c)
        if 0x1D400<=o<=0x1D419: out.append(chr(o-0x1D400+ord('A')))
        elif 0x1D41A<=o<=0x1D433: out.append(chr(o-0x1D41A+ord('a')))
        elif 0x1D7CE<=o<=0x1D7D7: out.append(chr(o-0x1D7CE+ord('0')))
        else: out.append(c)
    return "".join(out)

def _is_mathbold_char(c):
    o = ord(c)
    return (0x1D400<=o<=0x1D433) or (0x1D7CE<=o<=0x1D7D7)


def _detect_underline_answers(pdf_path: str) -> dict:
    try: reader = pypdf.PdfReader(pdf_path)
    except Exception: return {}
    results = {}
    for page in reader.pages:
        co = page.get('/Contents')
        if not co: continue
        try:
            if isinstance(co, pypdf.generic.ArrayObject):
                raw = b''.join(o.get_object().get_data() for o in co)
            else:
                raw = co.get_object().get_data()
        except Exception: continue
        s = raw.decode('latin-1', errors='replace')
        for m in re.finditer(r'(-?[\d.]+)\s+(-?[\d.]+)\s+(-?[\d.]+)\s+(-?[\d.]+)\s+re\s*\n\s*f', s):
            try: rw,rh = float(m.group(3)), float(m.group(4))
            except ValueError: continue
            if abs(rh)>=3.0 or abs(rw)<=3.0: continue
            before = s[:m.start()]
            cb = re.findall(r'\((.)\)\s+Tj', before)
            tb = ''.join(cb)
            choice = next((c.upper() for c in reversed(tb) if c.lower() in 'abcde'), None)
            if not choice: continue
            q_num = None
            for mq in re.finditer(r'(?<![.\d])(\d{1,2})\.', tb):
                n = int(mq.group(1))
                if 1<=n<=MAX_QUESTION_NUMBER: q_num=str(n)
            if q_num and q_num not in results:
                results[q_num] = choice
    return results


def parse_questions_from_pages(page_data: list, pdf_path: str = None) -> dict:
    """Returns {q_num: {prompt, choices, correct, omit, _norm, _tail}}.

    correct: comma-joined answer letter(s) e.g. "B" or "A,C"
    omit:    True if a red not-graded box was detected on this question
    """
    full_text_orig = ""
    full_bold: set = set()
    for plain, bold_idx in page_data:
        offset = len(full_text_orig)
        full_text_orig += plain + "\n"
        full_bold.update(i+offset for i in bold_idx)

    header_re = re.compile(r'[A-Z]{2,6}\s+\d{3,5}:?\s+.+?Page\s+\d+\s+of\s+\d+\s*', re.I)
    sto = {}; sc = []; op = 0
    for m in header_re.finditer(full_text_orig):
        while op < m.start():
            sto[len(sc)] = op; sc.append(full_text_orig[op]); op += 1
        sto[len(sc)] = m.start(); sc.append('\n'); op = m.end()
    while op < len(full_text_orig):
        sto[len(sc)] = op; sc.append(full_text_orig[op]); op += 1
    full_text = "".join(sc)

    o2s = {v:k for k,v in sto.items()}
    full_bold_s = set()
    for bi in full_bold:
        si = o2s.get(bi)
        if si is not None: full_bold_s.add(si)
    full_bold = full_bold_s
    full_text_for_bold = full_text

    mb_marks = []
    for _mb in re.finditer(r'([\U0001D400-\U0001D419])\.', full_text):
        mb_marks.append((_mb.start(), _mathbold_to_ascii(_mb.group(1))))
    if any(_is_mathbold_char(c) for c in full_text):
        full_text = _mathbold_to_ascii(full_text)
        full_text_for_bold = _mathbold_to_ascii(full_text_for_bold)

    # Strip cover-page instructions
    full_text = re.sub(
        r'(?:^|\n)([ \t]*)([1-9])\.\s+'
        r'((?:For\s+this\s+exam|[Uu]se\s+(?:ONLY|#2|only|a\s+#)|[Bb]ubble\s+in|'
        r'[Ww]hen\s+(?:filling|you\s+first)|[Yy]ou\s+(?:should|may\s+not|must)|'
        r'[Ww]e\s+will\s+(?:announce|periodically|collect)|'
        r'[Ii]f\s+you\s+(?:have\s+a\s+question|finish|are\s+taking)|'
        r'[Dd]o\s+not\s+(?:use|write|leave)|[Aa]fter\s+the\s+exam|'
        r'[Mm]ake\s+sure\s+you\s+have|[Ss]tudents\s+(?:are|must|should)|'
        r'[Tt]urn\s+off\s+|[Ss]ilence\s+your|[Pp]lease\s+(?:turn|write|put|silence)|'
        r'[Ee]njoy\s+your|I\.D\.\s*NUMBER|NAME\s*[:(]|TEST\s+FORM|'
        r'SUBJECT\s+SCORE|SECTION\s+(?:NUMBER|#))[^\n]{0,})',
        r'\n\1INSTRUCTION_STRIPPED', full_text, flags=re.I)

    # ── SUPERSCRIPT-EXPONENT FIX (pypdf-version-robust) ───────────────────────
    # Unit exponents like "mg L⁻¹", "s⁻¹", "M⁻¹" are extracted with a control
    # byte where the superscript was. Different pypdf versions place it
    # differently:
    #   pypdf 5.x:  "L\x001."      (no spaces)
    #   pypdf 6.x:  "L \x00 1."    (spaces around the control byte)
    # After the generic control-strip below, both collapse to "L 1." — which
    # looks EXACTLY like a new question marker ("1. ...") and makes the question
    # splitter swallow the rest of the question, dropping it entirely. This is
    # the single most common cause of a question going "missing" even though it
    # is plainly present in the source PDF.
    #
    # The reliable, version-independent signal is the CONTROL BYTE itself: a
    # control char (\x00-\x08, \x0b-\x1f) sitting between a letter and a
    # "digit." is always a mangled superscript exponent, never a real question
    # boundary (real question numbers are never preceded by a control byte).
    # We splice the bogus "digit." out and glue the exponent digit onto the
    # unit ("L\x001." or "L \x00 1." -> "L1"), so no phantom marker remains.
    # Because the trigger REQUIRES the control byte, this cannot touch any real
    # question number and is safe across all exam formats.
    full_text = re.sub(
        r'([A-Za-z])[ \t]*[\x00-\x08\x0b-\x1f][ \t]*([123])\.(?=[ \t]|[A-Za-z])',
        r'\1\2',
        full_text
    )

    full_text = re.sub(r'[\x00-\x08\x0b\x0c\x0f-\x11\x15-\x1f\x7f-\x9f]+',' ',full_text)
    full_text = re.sub(r'[\uFFFD\uE000-\uF8FF]+',' ',full_text)
    full_text = full_text.replace('\u2029',' ').replace('\u2028',' ')

    full_text = re.sub(
        r'(?:Short\s+Answer|Written\s+Answer|Free\s+Response)\s+Questions?\s+([1-9]\d?)\.',
        lambda m: f"\n{m.group(1)}.", full_text, flags=re.I)
    full_text = re.sub(r'\b([1-9]\d?)\.([ \t]*)([A-Z])',r'\1. \3',full_text)
    full_text = re.sub(
        r'(_{8,}[ \t]*)([1-9]\d?)([ \t]+)([A-Z][a-z])',
        lambda m: (f"{m.group(1)}{m.group(2)}. {m.group(4)}"
                   if 1<=int(m.group(2))<=MAX_QUESTION_NUMBER else m.group(0)),
        full_text)
    full_text = re.sub(
        r'(=|mass|weight|\()\s*(\d{1,3})\.\s+(\d{1,3})\s*(g\s*mol|g/mol|amu|kg|g\b|J\b|mL|cm|nm|atm|torr|mol)',
        lambda m: f"{m.group(1)} {m.group(2)}.{m.group(3)} {m.group(4)}",
        full_text, flags=re.I)
    full_text = re.sub(
        r'(\d\s*(?:V|kJ|mol|nm|pm|atm|torr|mL|L|°C)\b)( {1,2})([1-9]\d?)\.( {1,2})([A-Z][a-z]{2,})',
        lambda m: (f"{m.group(1)}\n{m.group(3)}. {m.group(5)}"
                   if 1<=int(m.group(3))<=MAX_QUESTION_NUMBER else m.group(0)),
        full_text)
    full_text = re.sub(r'\bof\s+\d+\s{1,4}(?=[1-9]\d?[.:][ \t])','\n',full_text)

    def _inl(m):
        val,sp1,qn,sp2,word = m.groups()
        if re.fullmatch(r'10\d?',val.strip()) or not(1<=int(qn)<=MAX_QUESTION_NUMBER):
            return m.group(0)
        return f"{val}\n{qn}. {word}"
    full_text = re.sub(
        r'(?<!\d{2})(?<![×x⇥⨯] \d)(?<![×x⇥⨯]\d)'
        r'(\b0?\.\d+|\b\d{1,2}\.?\d*)( {1,2})([1-9]\d?)\.( {1,2})([A-Z][a-z]{2,})',
        _inl, full_text)

    def _uwr(m):
        qn,word = m.group(1), m.group(2)
        if not(1<=int(qn)<=MAX_QUESTION_NUMBER): return m.group(0)
        full = m.group(0); ts = full.rfind(f"{qn}. {word}")
        return full[:ts] + " " + word
    full_text = re.sub(
        r'\d[\d.×x⇥⨯ \t]*(?:(?:[a-zA-Z]{1,3}[ \t]+){0,1}[a-zA-Z]{1,3}(?: {1,2})([1-9]\d?)\.(?: {1,2}))+([A-Z][a-z]{1,})',
        _uwr, full_text)

    pb  = len(re.findall(r'(?:^|\n|\s{2,})\(A\)\s+\S', full_text, re.M))
    pa  = len(re.findall(r'\bA\)\s{2,}', full_text))
    pal = len(re.findall(r'(?:^|\n|[ \t]{2,})([a-e])\)\s+\S', full_text, re.M))
    lc  = len(re.findall(r'\ba\.\s{2,}', full_text))
    ud  = len(re.findall(r'\bA\.\s+', full_text))
    pat = pa+pal
    if pb>=3:
        cre=re.compile(r'\(([A-Ea-e])\)\s*'); mfn=lambda l:f"({l.upper()})"
    elif pat>=3 and pat>=ud//2:
        cre=re.compile(r'\b([A-Ea-e])\)\s+')
        _pc='lower' if pal>pa else 'upper'
        mfn=lambda l,_c=_pc:(l.lower() if _c=='lower' else l.upper())+")"
    elif lc>=ud:
        cre=re.compile(r'\b([a-e])\.\s+'); mfn=lambda l:l.lower()+"."
    else:
        cre=re.compile(r'\b([A-E])\.\s+'); mfn=lambda l:f"{l}."

    _afmt = [
        (re.compile(r'\(([A-Ea-e])\)\s*'),  lambda l:f"({l.upper()})"),
        (re.compile(r'\b([A-E])\.\s+'),      lambda l:f"{l}."),
        (re.compile(r'\b([a-e])\.\s+'),      lambda l:f"{l.lower()}."),
        (re.compile(r'\b([A-Ea-e])\)\s+'),   lambda l:f"{l.upper()})"),
    ]
    _fb = [(rx,mk) for rx,mk in _afmt if rx.pattern!=cre.pattern]

    # Debug aid: set EXAM_MAPPER_DUMP_FULLTEXT=/path to dump the fully-normalised
    # text right before question-splitting. Useful for diagnosing exactly why a
    # specific question did or didn't split where expected.
    if os.environ.get("EXAM_MAPPER_DUMP_FULLTEXT") and pdf_path:
        try:
            with open(os.environ["EXAM_MAPPER_DUMP_FULLTEXT"], "w") as _df:
                _df.write(full_text)
        except Exception:
            pass

    raw  = Q_SPLIT_RE.split(full_text)
    qs   = {}
    cur  = 0
    maxq = 0
    _unattached_omits = []   # (q_num, note_excerpt) notes we could not confidently attach
    _ul  = _detect_underline_answers(pdf_path) if pdf_path else {}
    _rb  = _detect_red_boxes(pdf_path)         if pdf_path else {}

    for i in range(1, len(raw)-1, 2):
        q_num = raw[i].strip(); content = raw[i+1]
        qi = int(q_num)
        if qi < maxq-1: continue
        if qi > maxq: maxq = qi
        snip    = content[:30].strip()
        blkabs  = full_text_for_bold.find(snip, cur) if snip else cur
        if blkabs==-1: blkabs=cur
        cur = blkabs+1

        # Red-box / omit-note handling — decided AFTER we locate the choices
        # (below), so we can tell a note that genuinely belongs to THIS
        # question (appears at/after its A-E choices, e.g. a "not graded" note
        # printed under the answers) from a floating red-box note that got
        # dumped into this question's prompt but actually annotates a
        # different, visually-adjacent question. The latter must NOT create a
        # false omit; it is recorded for manual review instead.
        _rb_flag = _rb.get(q_num)
        _two_txt = bool(re.search(r'two\s+answers?\s+are\s+correct|both.*correct|two\s+correct\s+answers?',content,re.I))
        # Provisional flag for the no-choice / free-response branches below
        # (a red-box OMIT flag from geometry is always trustworthy; a bare
        # text note in a question with no choices is treated conservatively).
        omit_flag = (_rb_flag=='omit')
        _pending_omit = None  # (q_num, note_excerpt) when a note can't be
                              # confidently attached — surfaced for review

        def _nc(rx,txt):
            mm=rx.search(txt)
            if not mm: return 0
            return len(rx.split(txt[mm.start():])) // 2
        mf = cre.search(content); are=cre; amf=mfn
        pn = _nc(cre,content) if mf else 0
        if pn<2:
            bn=pn
            for fr,fm in _fb:
                n=_nc(fr,content)
                if n>bn: bn=n; mf=fr.search(content); are=fr; amf=fm

        if not mf:
            pc = re.sub(r'\s+',' ',content.strip())[:300]
            isq = (len(pc)>10 and len(pc)<400 and
                   not re.search(r'Potentially\s+Useful|USEFUL\s+INFO|Reference\s+Data|'
                                 r'CONTINUE\s+TO\s+NEXT\s+PAGE|Turn\s+over|Blank\s+page|'
                                 r'Formula\s+Sheet|Constants\s+and\s+Equations',pc,re.I) and
                   re.search(r'\?|calculate|determine|find|what|which|how|draw|sketch|'
                             r'explain|describe|identify|write|show|complete|select|'
                             r'predict|rank|compare|give|indicate|choose|state|list|'
                             r'balance|name|define|derive|evaluate|correct|incorrect|'
                             r'following|shown|below|above|structure|symbol|diagram|'
                             r'represent|assign|classify|arrange|order|match',pc,re.I))
            if isq:
                qs[q_num]={'prompt':pc,'choices':{},'correct':'Unknown',
                           'omit':omit_flag,'_norm':_norm(pc),'_tail':_tail(pc)}
            continue

        ph = content[:mf.start()]
        if (re.search(r'\(\s*\d+\s*pts?\s*\)',ph) or
                re.search(r'\([a-e]\)\s*\(\s*\d+\s*pts?\s*\)',content)):
            pc = re.sub(r'\s+',' ',content.strip())[:300]
            if 10<len(pc)<400:
                qs[q_num]={'prompt':pc,'choices':{},'correct':'Unknown',
                           'omit':omit_flag,'_norm':_norm(pc),'_tail':_tail(pc)}
            continue

        prompt = re.sub(r'\s+',' ',content[:mf.start()].strip()).strip()
        cp     = content[mf.start():]
        pts    = are.split(cp)
        choices: dict = {}; cs: set = set(); cao = {}; asf = blkabs+mf.start()

        for j in range(1, len(pts)-1, 2):
            lr=pts[j].strip(); L=lr.upper(); rt=pts[j+1]
            ct=rt.split('\n')[0].strip()
            mk=amf(lr); ap=full_text_for_bold.find(mk,asf)
            if ap>=0: cao[L]=ap; asf=ap+len(mk)
            if _has_tick(lr) or _has_tick(ct) or _has_tick(rt[:100]):
                cs.add(L); ct=_strip_tick(ct)
            choices[L]=ct

        if not cs and cao: cs|=_detect_bold_answers(cao, full_bold)
        if not cs and cao and mb_marks:
            rs=min(cao.values()); re_=max(cao.values())+120
            for _p,_l in mb_marks:
                if rs<=_p<=re_ and _l in choices: cs.add(_l)
        if not cs and q_num in _ul:
            ul=_ul[q_num]
            if ul in choices: cs.add(ul)
        if not cs:
            ma=re.search(r'(?:Answer|ANSWER)\s*[=:]\s*([A-Ea-e](?:\s*(?:,|/|and|&)\s*[A-Ea-e])*)\b',content)
            if ma:
                for al in re.findall(r'[A-Ea-e]',ma.group(1)):
                    AL=al.upper()
                    if AL in choices: cs.add(AL)

        _ocr = bool(pdf_path) and '-ocr' in os.path.basename(pdf_path).lower()
        if not cs and choices and full_bold and not _ocr:
            bs=blkabs; nxt=None
            try:
                ni=int(q_num)+1
                mn=re.search(rf'(?<![0-9]){ni}\.\s',full_text_for_bold[bs+3:])
                if mn: nxt=bs+3+mn.start()
            except Exception: pass
            he=min(bs+len(content)+40,len(full_text_for_bold))
            be=min(nxt,he) if nxt else he
            mp={}; sf=bs
            for L in sorted(choices.keys()):
                mm=re.search(rf'(?<![A-Za-z0-9])[{L}{L.lower()}][.)]',full_text_for_bold[sf:be])
                if mm: p=sf+mm.start(); mp[L]=p; sf=p+1
            if len(mp)>=2:
                od=sorted(mp.items(),key=lambda kv:kv[1])
                sc2={}; mbm={}
                for k,(L,p) in enumerate(od):
                    e_=od[k+1][1] if k+1<len(od) else min(p+70,be)
                    mb_=sum(1 for ii in range(p,min(p+2,e_)) if ii in full_bold)
                    tb_=sum(1 for ii in range(p,e_)          if ii in full_bold)
                    mbm[L]=mb_; sc2[L]=mb_*3+tb_
                if any(sc2.values()):
                    rl=sorted(sc2.items(),key=lambda kv:kv[1],reverse=True)
                    best,bs2=rl[0]; run=rl[1][1] if len(rl)>1 else 0
                    bm=mbm[best]; bt=bs2-bm*3
                    ombs=sum(1 for L,mb in mbm.items() if L!=best and mb>=1)
                    if bm>=1 and ombs==0 and bs2>run: cs.add(best)
                    elif bt>=4 and bs2>run*2: cs.add(best)
                    for L,mb in mbm.items():
                        if mb>=1 and L!=best: cs.add(L)

        for L in list(choices.keys()):
            t=choices[L]
            for stop in ('END OF QUESTIONS','Potentially Use','CHEM 1311','CHEM 1312',
                         'electron:  mass','proton:  mass','neutron:  mass',
                         'Avogadro','speed of light','Planck',
                         '=  –1.60','= 1.673','= 1.675','two answers are correct'):
                ix=t.find(stop)
                if ix>=0: choices[L]=t[:ix].strip()
            choices[L]=re.sub(r'\s*_{5,}\s*$','',choices[L]).strip()

        correct = ",".join(sorted(cs)) if cs else "Unknown"
        if choices:
            # Decide omit by WHERE the note sits. A "not graded / credit to all"
            # note that appears at or after this question's A-E choices belongs
            # to this question (e.g. a note printed beneath the answers). A note
            # found ONLY in the prompt — before any choice — is a red box that
            # floated in from a visually-adjacent question and must not create a
            # false omit here; it is recorded for manual review instead.
            _prompt_part = content[:mf.start()]
            _choice_part = content[mf.start():]
            _cm = OMIT_PHRASES.search(_choice_part)
            _note_in_prompt  = bool(OMIT_PHRASES.search(_prompt_part))
            # A note only auto-omits THIS question when it sits in the choices
            # region, is NOT past an end-of-exam boundary, and does NOT contain
            # cross-reference language showing it actually describes a DIFFERENT
            # question/form ("on other versions", "corresponding question", "no
            # salt bridge", "mistake in the line notation", etc.). Those float
            # in reading order and would otherwise create false omits on the
            # wrong master row, so they are recorded for manual review instead.
            _CROSSREF = re.compile(
                r'other\s+version|corresponding\s+question|on\s+other\s+form|'
                r'different\s+version|no\s+salt\s+bridge|answer\s+is\s+not\s+there|'
                r'mistake\s+in\s+the\s+line|has\s+a\s+mistake|on\s+another\s+form',
                re.I)
            _note_in_choices = False
            if _cm:
                _win = _choice_part[max(0,_cm.start()-150):_cm.end()+150]
                _between = _choice_part[:_cm.start()]
                _boundary = 'END OF QUESTIONS' in _between.upper()
                _crossref = bool(_CROSSREF.search(_win)) or bool(_CROSSREF.search(_prompt_part))
                _note_in_choices = (not _boundary) and (not _crossref)
            # Also disqualify a prompt-region note that carries cross-ref text.
            if _note_in_prompt and _CROSSREF.search(_prompt_part):
                _note_in_prompt = False
                _prompt_note_floating = True
            else:
                _prompt_note_floating = False
            if _rb_flag=='omit':
                omit_flag = True
            elif _note_in_choices and not _two_txt:
                omit_flag = True
            elif (bool(OMIT_PHRASES.search(_prompt_part)) or _cm) and not _two_txt:
                omit_flag = False
                _src = _prompt_part if OMIT_PHRASES.search(_prompt_part) else _choice_part
                _m = OMIT_PHRASES.search(_src)
                _pending_omit = (q_num, _src[max(0,_m.start()-10):_m.start()+70].strip())
            qs[q_num]={'prompt':prompt,'choices':choices,'correct':correct,
                       'omit':omit_flag,'_norm':_norm(prompt),'_tail':_tail(prompt)}
            if _pending_omit:
                _unattached_omits.append(_pending_omit)

    # Continuity check
    if qs:
        fn2=sorted(int(k) for k in qs)
        gaps=[i for i in range(1,fn2[-1]+1) if i not in set(fn2)]
        jumps=[(fn2[i-1],fn2[i]) for i in range(1,len(fn2)) if fn2[i]!=fn2[i-1]+1]
        lbl=os.path.basename(pdf_path) if pdf_path else "<unnamed>"
        if gaps:  _logger.warning("%s GAPS: %s",lbl,gaps)
        if jumps: _logger.warning("%s JUMPS: %s",lbl,"; ".join(f"{a} then {b}" for a,b in jumps))
    # Record any omit notes we could not confidently attach to a specific
    # question in a side-channel keyed by file, so the caller can flag them for
    # manual review rather than silently dropping a real "not graded" note.
    if pdf_path and _unattached_omits:
        _UNATTACHED_OMITS[os.path.basename(pdf_path)] = _unattached_omits
    return qs


# ── FUZZY MATCHING ────────────────────────────────────────────────────────────

def choice_sim(a,b):
    na=re.sub(r'[^a-z0-9.]','',a.lower()).strip('.')
    nb=re.sub(r'[^a-z0-9.]','',b.lower()).strip('.')
    if not na or not nb: return 0.0
    return _sim(na,nb)

def _choice_token_set(choices):
    BOIL=('alloftheabove','noneoftheabove','allareequal','alloftheaboveareequal',
          'allequal','bothaandb','noneofthese','aandb','allofthese')
    toks=set()
    for t in choices.values():
        n=re.sub(r'[^a-z0-9]','',str(t).lower())
        if n and n not in BOIL: toks.add(n)
    return frozenset(toks)

def _choice_overlap(ca,cb):
    sa,sb=_choice_token_set(ca),_choice_token_set(cb)
    if not sa or not sb: return 1.0
    return len(sa&sb)/len(sa|sb)

def map_choice_order(rc,tc):
    seq=[]
    for rl in sorted(rc.keys()):
        rt=rc.get(rl,"")
        if not rt: seq.append("?"); continue
        bt,bs="?",0.0
        for tl,tt in tc.items():
            s=choice_sim(rt,tt)
            if s>bs: bs=s; bt=tl
        seq.append(bt if bs>=0.70 else "?")
    return " -> ".join(seq)

def translate_answer(ref,order):
    if ref in ('Unknown','N/A','',None): return 'Unknown'
    parts=  [p.strip() for p in order.split('->')]
    rl=['A','B','C','D','E'][:len(parts)]
    out=set()
    for L in str(ref).split(','):
        L=L.strip().upper()
        try: out.add(parts[rl.index(L)])
        except (ValueError,IndexError): pass
    return ",".join(sorted(out)) if out else 'Unknown'


# ── VARIANT GROUPING ──────────────────────────────────────────────────────────

def _is_same_concept(dx,dy):
    fs=_sim(dx['_norm'],dy['_norm']); ws=_word_overlap(dx['prompt'],dy['prompt'])
    if not (fs>=MATCH_THRESHOLD or ws>=WORD_OVERLAP_THRESHOLD): return False,False
    if _choice_overlap(dx.get('choices',{}),dy.get('choices',{}))<CHOICE_OVERLAP_MIN:
        return True,True
    return True, fs<VARIANT_THRESHOLD

def _concept_key(d):
    sig="|".join(sorted(_choice_token_set(d.get('choices',{}))))
    return d['_norm']+" ##CH## "+sig

def _pair_score(a,b):
    return max(_sim(a['_norm'],b['_norm']),_word_overlap(a['prompt'],b['prompt']))


def _global_match_pair(anchor_qs: dict, tgt_qs: dict) -> dict:
    """Global optimal one-to-one matching — prevents missed questions."""
    thr = min(MATCH_THRESHOLD, WORD_OVERLAP_THRESHOLD)
    cands = []
    for an,ad in anchor_qs.items():
        for tn,td in tgt_qs.items():
            s=_pair_score(ad,td)
            if s>=thr: cands.append((s,an,tn))
    cands.sort(key=lambda c:c[0],reverse=True)
    ua=set(); ut=set(); res={}
    for s,an,tn in cands:
        if an in ua or tn in ut: continue
        res[an]=(tn,s); ua.add(an); ut.add(tn)
    return res


def group_variants(form_data: dict, anchor_form: str) -> list:
    """Cross-match with global matching + union orphan pass + OMIT propagation."""
    aq=form_data[anchor_form]; of=[f for f in form_data if f!=anchor_form]
    variants=[]; master_num=0
    gm={f:_global_match_pair(aq,form_data[f]) for f in of}
    claimed={f:{t for(t,_)in gm[f].values()} for f in of}

    for qna,da in aq.items():
        master_num+=1
        fm={}
        for f in of:
            me=gm[f].get(qna)
            if me:
                bn,bs=me; md=form_data[f][bn]
                sc,iv=_is_same_concept(da,md)
                if sc: fm[f]={'q_num':bn,'data':md,'is_variant':iv,'score':bs}
                else:  fm[f]=None; claimed[f].discard(bn)
            else: fm[f]=None

        ak=_concept_key(da); k2b={ak:0}; nb=1
        vb=defaultdict(lambda:{'forms':[],'prompt':'','choices':{},'correct':'Unknown','omit':False,'form_data':{}})
        vb[0]['forms'].append(anchor_form)
        vb[0]['prompt']=da['prompt']; vb[0]['choices']=da['choices']
        vb[0]['correct']=da['correct']; vb[0]['omit']=da.get('omit',False)
        vb[0]['form_data'][anchor_form]={'q_num':qna,'choices':da['choices'],
            'correct':da['correct'],'omit':da.get('omit',False),
            'choice_order':" -> ".join(sorted(da['choices'].keys()))}

        for f in of:
            match=fm[f]
            if match is None:
                vb[0]['form_data'][f]={'q_num':'MISSING','choices':{},'correct':'N/A','omit':False,'choice_order':'N/A'}
                continue
            md=match['data']; mk=_concept_key(md)
            if not match['is_variant']: bi=0; rc=da['choices']
            else:
                bi=k2b.get(mk)
                if bi is None: bi=nb; k2b[mk]=bi; nb+=1
                rc=vb[bi]['choices'] or md['choices']
            vb[bi]['forms'].append(f)
            co=(" -> ".join(sorted(md['choices'].keys()))
                if not vb[bi]['choices'] or rc is md['choices']
                else map_choice_order(rc,md['choices']))
            tc=(md['correct'] if md['correct']!='Unknown'
                else translate_answer(da['correct'],co)
                if da['correct']!='Unknown' and not match['is_variant']
                else md['correct'])
            vb[bi]['form_data'][f]={'q_num':match['q_num'],'choices':md['choices'],
                'correct':tc,'omit':md.get('omit',False),'choice_order':co}
            if bi>0 and not vb[bi]['prompt']:
                vb[bi]['prompt']=md['prompt']; vb[bi]['choices']=md['choices']
                vb[bi]['correct']=md['correct']; vb[bi]['omit']=md.get('omit',False)
                vb[bi]['form_data'][f]['choice_order']=" -> ".join(sorted(md['choices'].keys()))

        for bi in sorted(vb.keys()):
            b=vb[bi]
            variants.append({'master_num':master_num,'variant_idx':bi+1,
                'variant_forms':sorted(b['forms']),'prompt':b['prompt'],
                'choices':b['choices'],'correct':b['correct'],'omit':b['omit'],
                'form_data':b['form_data']})

    # Pass 2: orphans
    orphans=[]
    for f in of:
        for qn,d in form_data[f].items():
            if qn not in claimed[f]: orphans.append((f,qn,d))
    n=len(orphans); pc2=[]
    for i in range(n):
        fi,qi,di=orphans[i]
        for j in range(i+1,n):
            fj,qj,dj=orphans[j]
            if fj==fi: continue
            sc,_=_is_same_concept(di,dj)
            if sc: pc2.append((_pair_score(di,dj),i,j))
    pc2.sort(key=lambda c:c[0],reverse=True)
    co2={}; cl2={}; nc2=0
    for s,i,j in pc2:
        fi,qi,di=orphans[i]; fj,qj,dj=orphans[j]
        ci,cj=co2.get(i),co2.get(j)
        if ci is None and cj is None:
            cid=nc2; nc2+=1; cl2[cid]={fi:(qi,di),fj:(qj,dj)}; co2[i]=cid; co2[j]=cid
        elif ci is not None and cj is None:
            if fj not in cl2[ci]: cl2[ci][fj]=(qj,dj); co2[j]=ci
        elif cj is not None and ci is None:
            if fi not in cl2[cj]: cl2[cj][fi]=(qi,di); co2[i]=cj
    for i,(fi,qi,di) in enumerate(orphans):
        if i not in co2: cid=nc2; nc2+=1; cl2[cid]={fi:(qi,di)}; co2[i]=cid
    for cid in sorted(cl2.keys()):
        cl=cl2[cid]; master_num+=1
        rf=sorted(cl.keys(),key=lambda f:(FORMS.index(f) if f in FORMS else 99,f))[0]
        rq,rd=cl[rf]
        fdo={anchor_form:{'q_num':'MISSING','choices':{},'correct':'N/A','omit':False,'choice_order':'N/A'}}
        for f in of:
            if f==rf:
                fdo[f]={'q_num':rq,'choices':rd['choices'],'correct':rd['correct'],
                         'omit':rd.get('omit',False),'choice_order':" -> ".join(sorted(rd['choices'].keys()))}
            elif f in cl:
                qf,df=cl[f]; co_=map_choice_order(rd['choices'],df['choices'])
                tc_=(df['correct'] if df['correct']!='Unknown'
                     else translate_answer(rd['correct'],co_) if rd['correct']!='Unknown'
                     else 'Unknown')
                fdo[f]={'q_num':qf,'choices':df['choices'],'correct':tc_,
                         'omit':df.get('omit',False),'choice_order':co_}
            else:
                fdo[f]={'q_num':'MISSING','choices':{},'correct':'N/A','omit':False,'choice_order':'N/A'}
        variants.append({'master_num':master_num,'variant_idx':1,'variant_forms':sorted(cl.keys()),
            'prompt':rd['prompt'],'choices':rd['choices'],'correct':rd['correct'],
            'omit':rd.get('omit',False),'form_data':fdo})
    return variants


# ── BUILD DATAFRAME ───────────────────────────────────────────────────────────

def build_mapping_dataframe(form_data: dict, meta: dict,
                             answer_key: dict = None) -> pd.DataFrame:
    def _fsk(f):
        try: return (0,FORMS.index(f))
        except ValueError:
            try: return (1,int(f))
            except ValueError: return (2,f)
    avf=sorted(form_data.keys(),key=_fsk); af=avf[0] if avf else None
    if not af: return pd.DataFrame()
    if answer_key:
        for mn,(qn,d) in enumerate(form_data[af].items(),1):
            if mn in answer_key: d['correct']=answer_key[mn]

    variants=group_variants(form_data,af)

    # ── COMPLETENESS GUARD ────────────────────────────────────────────────────
    # Every parsed question from every form MUST appear in the output.
    # If grouping missed any, force them in as explicit rows.
    seen=set()
    for v in variants:
        for f,fd in v['form_data'].items():
            qn=fd.get('q_num')
            if qn not in ('MISSING','N/A',None): seen.add((f,str(qn)))
    miss=[(f,qn) for f,qs in form_data.items() for qn in qs if (f,str(qn)) not in seen]
    if miss:
        nm=max((v['master_num'] for v in variants),default=0)
        miss.sort(key=lambda p:(p[0],int(p[1]) if p[1].isdigit() else 9999))
        for f,qn in miss:
            d=form_data[f][qn]; nm+=1
            fdo={ff:({'q_num':qn,'choices':d.get('choices',{}),'correct':d.get('correct','Unknown'),
                      'omit':d.get('omit',False),'choice_order':" -> ".join(sorted(d.get('choices',{}).keys()))}
                     if ff==f else {'q_num':'MISSING','choices':{},'correct':'N/A','omit':False,'choice_order':'N/A'})
                 for ff in form_data}
            variants.append({'master_num':nm,'variant_idx':1,'variant_forms':[f],
                'prompt':d.get('prompt',''),'choices':d.get('choices',{}),'correct':d.get('correct','Unknown'),
                'omit':d.get('omit',False),'form_data':fdo})
        print(f"  [completeness guard] recovered {len(miss)} question(s) not reached by grouping.")

    # Completeness guard
    seen=set()
    for v in variants:
        for f,fd in v['form_data'].items():
            qn=fd.get('q_num')
            if qn not in ('MISSING','N/A',None): seen.add((f,str(qn)))
    miss=[(f,qn) for f,qs in form_data.items() for qn in qs if (f,str(qn)) not in seen]
    if miss:
        nm=max((v['master_num'] for v in variants),default=0)
        miss.sort(key=lambda p:(p[0],int(p[1]) if p[1].isdigit() else 9999))
        for f,qn in miss:
            d=form_data[f][qn]; nm+=1
            fdo={ff:({'q_num':qn,'choices':d.get('choices',{}),'correct':d.get('correct','Unknown'),
                      'omit':d.get('omit',False),'choice_order':" -> ".join(sorted(d.get('choices',{}).keys()))}
                     if ff==f else {'q_num':'MISSING','choices':{},'correct':'N/A','omit':False,'choice_order':'N/A'})
                 for ff in form_data}
            variants.append({'master_num':nm,'variant_idx':1,'variant_forms':[f],
                'prompt':d.get('prompt',''),'choices':d.get('choices',{}),'correct':d.get('correct','Unknown'),
                'omit':d.get('omit',False),'form_data':fdo})
        print(f"  [completeness] recovered {len(miss)} question(s).")

    rows=[]
    for v in variants:
        mn=v['master_num']; vi=v['variant_idx']; vf=v['variant_forms']
        tv=sum(1 for vv in variants if vv['master_num']==mn)
        vlbl=f"V{vi} ({'/'.join(vf)})" if tv>1 else ""
        mc=v['choices']
        # OMIT belongs to THIS master row only. One master row represents one
        # question across all forms (Form A #2 == Form B #7 == ... are the same
        # row), so if that question was flagged not-graded on any of its forms,
        # this single row is OMIT="Yes" — and no OTHER master row is affected.
        any_omit=v.get('omit',False) or any(
            fd.get('omit',False) for fd in v['form_data'].values()
            if fd.get('q_num') not in ('MISSING','N/A'))
        row={"OMIT":"Yes" if any_omit else "","Semester":meta.get("semester",""),
             "Course":meta.get("course",""),"Exam":meta.get("exam",""),
             "Master Question Number":mn,"Variant":vlbl,"Question Prompt":v['prompt']}
        for L in ['A','B','C','D','E']: row[f"Choice {L}"]=mc.get(L,"")
        for f in avf:
            pf=f"Form {f}"; fd=v['form_data'].get(f)
            if fd is None:
                row[f"{pf} #"]="N/A"; row[f"{pf} Choice Order"]="N/A"; row[f"{pf} Correct Choice"]="N/A"
            elif fd['q_num']=='MISSING':
                row[f"{pf} #"]="MISSING"; row[f"{pf} Choice Order"]="N/A"; row[f"{pf} Correct Choice"]="N/A"
            else:
                row[f"{pf} #"]=fd['q_num']; row[f"{pf} Choice Order"]=fd['choice_order']
                row[f"{pf} Correct Choice"]=fd['correct']
        rows.append(row)
    return pd.DataFrame(rows)


# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────

HEADER_FONT=Font(name="Arial",bold=True,color="FFFFFF",size=10)
DATA_FONT  =Font(name="Arial",size=10)
THIN       =Side(style="thin",color="BFBFBF")
THIN_BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
BASE_FILL  =PatternFill("solid",fgColor="1F4E79")
FORM_HDR={"A":PatternFill("solid",fgColor="2E75B6"),"B":PatternFill("solid",fgColor="375623"),
          "C":PatternFill("solid",fgColor="7F6000"),"D":PatternFill("solid",fgColor="843C0C"),
          "E":PatternFill("solid",fgColor="4B2472"),"F":PatternFill("solid",fgColor="1E5750"),
          "G":PatternFill("solid",fgColor="843C8C"),"H":PatternFill("solid",fgColor="0E4A6E")}
FORM_ROW={"A":PatternFill("solid",fgColor="DEEAF1"),"B":PatternFill("solid",fgColor="E2EFDA"),
          "C":PatternFill("solid",fgColor="FFF2CC"),"D":PatternFill("solid",fgColor="FCE4D6"),
          "E":PatternFill("solid",fgColor="EDE7F6"),"F":PatternFill("solid",fgColor="E0F7FA"),
          "G":PatternFill("solid",fgColor="F3E5F5"),"H":PatternFill("solid",fgColor="E1F0F8")}
VARIANT_FILL={1:PatternFill("solid",fgColor="EBF3FB"),2:PatternFill("solid",fgColor="FEF9EC"),
              3:PatternFill("solid",fgColor="EDFAED"),4:PatternFill("solid",fgColor="FDEEF0")}
VARIANT_FONT_COLOR={1:"1F4E79",2:"7F4F00",3:"1E5C1E",4:"7B0D1E"}
OMIT_FILL=PatternFill("solid",fgColor="C00000")
OMIT_FONT=Font(name="Arial",bold=True,size=10,color="FFFFFF")
COL_WIDTHS={"OMIT":6,"Semester":14,"Course":12,"Exam":10,"Master Question Number":8,
            "Variant":14,"Question Prompt":52,**{f"Choice {l}":28 for l in "ABCDE"}}

def _clean(v):
    if isinstance(v,str): return re.sub(r'[\x00-\x1f\x7f-\x9f]','',v.replace('ℓ','l'))
    return v

def save_to_excel(df: pd.DataFrame, out_path: str):
    for col in EXTRA_COLS:
        if col not in df.columns: df[col]=""
    for col in df.columns: df[col]=df[col].apply(_clean)
    df.to_excel(out_path,index=False,engine="openpyxl")
    wb=load_workbook(out_path); ws=wb.active; ws.freeze_panes="A2"
    cols=list(df.columns); n=len(df)+1
    omit_ci=cols.index("OMIT")+1
    for ci,cn in enumerate(cols,1):
        cell=ws.cell(row=1,column=ci)
        fill=BASE_FILL
        for f in FORMS:
            if cn.startswith(f"Form {f}"):
                fill=FORM_HDR.get(f,PatternFill("solid",fgColor="595959")); break
        else:
            if cn in EXTRA_COLS: fill=PatternFill("solid",fgColor="D9E1F2")
            elif cn=="Variant":  fill=PatternFill("solid",fgColor="5B4A9C")
            elif cn=="OMIT":     fill=PatternFill("solid",fgColor="C00000")
        cell.font=(HEADER_FONT if cn not in EXTRA_COLS
                   else Font(name="Arial",bold=True,size=10,color="000000"))
        cell.fill=fill; cell.border=THIN_BORDER
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

    for ri in range(2,n+1):
        vv=ws.cell(row=ri,column=cols.index("Variant")+1).value or ""
        vm=re.match(r'V(\d+)',str(vv)); vn=int(vm.group(1)) if vm else 1
        ov=str(ws.cell(row=ri,column=omit_ci).value or "")
        for ci,cn in enumerate(cols,1):
            cell=ws.cell(row=ri,column=ci)
            cell.font=DATA_FONT; cell.border=THIN_BORDER
            cell.alignment=Alignment(vertical="center",wrap_text=False)
            applied=False
            for f in FORMS:
                if cn.startswith(f"Form {f}"):
                    cell.fill=FORM_ROW.get(f,PatternFill("solid",fgColor="EFEFEF")); applied=True; break
            if not applied and cn=="Variant" and vv:
                cell.fill=VARIANT_FILL.get(vn,PatternFill("solid",fgColor="FFFFFF"))
                cell.font=Font(name="Arial",bold=True,size=10,color=VARIANT_FONT_COLOR.get(vn,"000000"))
                cell.alignment=Alignment(horizontal="center",vertical="center")
            if cn=="OMIT" and ov.strip().upper()=="YES":
                cell.fill=OMIT_FILL; cell.font=OMIT_FONT
                cell.alignment=Alignment(horizontal="center",vertical="center")
            val=str(cell.value or "")
            if val=="MISSING":
                cell.fill=PatternFill("solid",fgColor="FADADD")
                cell.font=Font(name="Arial",bold=True,size=10,color="C00000")
            elif val=="Unknown":
                cell.fill=PatternFill("solid",fgColor="FFF2CC")
                cell.font=Font(name="Arial",size=10,color="BF8F00")

    for cn in cols:
        cl=get_column_letter(cols.index(cn)+1)
        if cn in COL_WIDTHS: ws.column_dimensions[cl].width=COL_WIDTHS[cn]
        elif cn in EXTRA_COLS: ws.column_dimensions[cl].width=10
        elif "Choice Order" in cn: ws.column_dimensions[cl].width=26
        elif "Correct Choice" in cn: ws.column_dimensions[cl].width=14
        elif cn.endswith(" #"): ws.column_dimensions[cl].width=10
        else: ws.column_dimensions[cl].width=14
    ws.row_dimensions[1].height=36
    for r in range(2,n+1): ws.row_dimensions[r].height=18
    wb.save(out_path)
    print(f"  Saved -> {out_path}")


# ── DIRECTORY WALKER ──────────────────────────────────────────────────────────

def _classify_form(fname):
    base=os.path.splitext(os.path.basename(fname))[0]
    L2N=lambda c:ord(c.upper())-ord('A')
    m=re.search(r'(?<![A-Za-z])v?EARLY[\s_-]+([A-H])\b',base,re.I)
    if m: return (1,L2N(m.group(1)),f"early{m.group(1).upper()}")
    m=re.search(r'(?<![A-Za-z])v?LATE[\s_-]+([A-H])\b',base,re.I)
    if m: return (2,L2N(m.group(1)),f"late{m.group(1).upper()}")
    m=re.search(r'(?<![A-Za-z])v([A-H])(?![A-Za-z])',base,re.I)
    if m: return (0,L2N(m.group(1)),f"v{m.group(1).upper()}")
    if re.search(r'(?<![A-Za-z])v?EARLY(?![A-Za-z])',base,re.I): return (1,0,"early")
    if re.search(r'(?<![A-Za-z])v?LATE(?![A-Za-z])',base,re.I):  return (2,0,"late")
    m=re.search(r'(?<![A-Za-z])v0*([1-9]\d?)(?:[^A-Za-z0-9]|$)',base,re.I)
    if m: return (0,int(m.group(1))-1,f"v{int(m.group(1)):03d}")
    m=re.search(r'(?:Exam|exam)\s*\d*[-\s_]*([A-H])(?:[-_\s\d]|$)',base,re.I)
    if m: return (0,L2N(m.group(1)),f"exam{m.group(1).upper()}")
    m=re.search(r'\bform[_\-\s]*([A-H])\b',base,re.I)
    if m: return (0,L2N(m.group(1)),f"form{m.group(1).upper()}")
    m=re.search(r'\bform([A-H])(?:[^A-Za-z]|$)',base,re.I)
    if m: return (0,L2N(m.group(1)),f"form{m.group(1).upper()}")
    m=re.search(r'\b(?:exam|key|midterm|final|quiz|test|section)\w*?[_\-\s]*\d*[_\-\s]*([A-H])(?:[_\-\s]|$)',base,re.I)
    if m: return (0,L2N(m.group(1)),f"kw{m.group(1).upper()}")
    ai=list(re.finditer(r'(?:^|[^A-Za-z])([A-H])(?:[^A-Za-z]|$)',base))
    if ai:
        L=ai[-1].group(1).upper(); return (0,L2N(L),f"iso{L}")
    stripped=re.sub(r'[_\-\s]*(key|ocr|ans|answer|answers|final|copy|version|ver)\s*$','',base,flags=re.I)
    m=re.search(r'([A-H])$',stripped,re.I)
    if m: L=m.group(1).upper(); return (0,L2N(L),f"trail{L}")
    m=re.search(r'[-_]0*([1-9]\d?)$',base)
    if m: return (3,int(m.group(1))-1,f"num{int(m.group(1))}")
    return None

def _form_from_filename(fname):
    c=_classify_form(fname)
    if c is None: return None
    cat,rank,_=c
    if cat==0: return chr(ord('A')+rank) if rank<26 else str(rank+1)
    if cat==3: return str(rank+1)
    return chr(ord('A')+rank)

def assign_form_letters(filenames):
    classified=[]; unrecognised=[]
    for f in filenames:
        c=_classify_form(f)
        if c is None: unrecognised.append(f)
        else: classified.append((c[0],c[1],c[2],f))
    classified.sort(key=lambda x:(x[0],x[1],x[2],x[3]))
    def _ltr(i): return chr(ord('A')+i) if i<26 else chr(ord('A')+i//26-1)+chr(ord('A')+i%26)
    mapping={}; used=set()
    for cat,rank,raw,f in classified:
        if cat==0 and rank<26 and _ltr(rank) not in used:
            let=_ltr(rank); mapping[f]=let; used.add(let)
    for cat,rank,raw,f in classified:
        if f in mapping: continue
        i=0
        while _ltr(i) in used: i+=1
        let=_ltr(i); mapping[f]=let; used.add(let)
    return dict(sorted(mapping.items(),key=lambda kv:(len(kv[1]),kv[1]))), unrecognised


def process_pdfs(pdf_map,meta_override,answer_key=None,out_path="Exam_Mapping.xlsx",log=None):
    form_data={}; shared_meta={}
    for form in sorted(pdf_map.keys()):
        fpath=pdf_map[form]; fname=os.path.basename(fpath)
        entry={"file":fname,"form":form,"status":"OK","questions":0,"unknown":0,
               "gaps":[],"jumps":[],"semester":"","exam":"","warnings":[],"errors":[]}
        print(f"  Parsing Form {form}: {fname}")
        try:
            pd_=extract_page_data(fpath); pt=extract_text(fpath)
            if not shared_meta:
                shared_meta=extract_metadata(pt,fpath)
                shared_meta.update({k:v for k,v in meta_override.items() if v})
            qs=parse_questions_from_pages(pd_,pdf_path=fpath)
            found=sorted(int(k) for k in qs)
            unknown=[k for k,d in qs.items() if d['correct']=='Unknown']
            gaps=([i for i in range(1,max(found)+1) if i not in set(found)] if found else [])
            jumps=([(found[i-1],found[i]) for i in range(1,len(found)) if found[i]!=found[i-1]+1] if found else [])
            # Image-likely: choices are images (empty/blank), OR the prompt asks
            # the student to pick among visual answers (Lewis structures,
            # diagrams, graphs, structures) which are not readable as text.
            _IMG_PROMPT = re.compile(
                r'lewis\s+(?:structure|symbol|dot)|'
                r'which\s+(?:of\s+the\s+following\s+)?(?:structure|diagram|graph|'
                r'orbital|molecule|drawing|figure|picture|shape)|'
                r'select\s+the\s+(?:correct\s+|best\s+)?(?:structure|lewis|diagram|'
                r'orbital\s+diagram|molecular)|'
                r'best\s+(?:lewis\s+)?(?:structure|description)|'
                r'(?:molecular|electron[- ]dot|orbital)\s+(?:structure|diagram|geometry\s+shown)|'
                r'shown\s+(?:below|above|in\s+the\s+figure)',
                re.I)
            def _img_like(k):
                ch = qs[k]['choices']
                if ch and all(not str(v).strip() for v in ch.values()):
                    return True   # has A-E slots but all blank -> image options
                if not ch and _IMG_PROMPT.search(qs[k]['prompt']):
                    return True   # no choices + visual-answer wording
                return False
            img_u=sorted((k for k in unknown if _img_like(k)),key=int)
            fr_u =sorted((k for k in unknown if not qs[k]['choices'] and not _img_like(k)),key=int)
            mk_u =sorted((k for k in unknown if qs[k]['choices'] and any(str(v).strip() for v in qs[k]['choices'].values())),key=int)
            entry.update({"questions":len(qs),"unknown":len(unknown),"gaps":gaps,"jumps":jumps,
                          "mark_unknown":mk_u,"img_unknown":img_u,"fr_unknown":fr_u,
                          "review_omits":_UNATTACHED_OMITS.get(os.path.basename(fpath),[]),
                          "semester":shared_meta.get("semester",""),"exam":shared_meta.get("exam","")})
            if not shared_meta.get("semester"): entry["warnings"].append("Semester not detected in PDF header")
            if gaps:  entry["warnings"].append(f"Missing question numbers: {gaps}")
            if jumps: entry["warnings"].append("NUMBER SEQUENCE BREAK — questions jump: "+"; ".join(f"{a} then {b}" for a,b in jumps))
            if mk_u:  entry["warnings"].append(f"{len(mk_u)} MC question(s) with text choices but no detectable answer mark (key may be unmarked): Q{', Q'.join(mk_u)}")
            if img_u: entry["warnings"].append(f"{len(img_u)} possible image/diagram question(s) — answer choices could not be read, needs manual review: Q{', Q'.join(img_u)}")
            if fr_u:  entry["warnings"].append(f"{len(fr_u)} free-response question(s) (no MC choices, expected): Q{', Q'.join(fr_u)}")
            if len(qs)==0:  entry["status"]="EMPTY"; entry["errors"].append("No questions parsed — check PDF format or encoding")
            elif gaps or jumps or img_u or len(mk_u)>2: entry["status"]="CHECK"
            print(f"    -> {len(qs)} questions  gaps={gaps}  jumps={jumps}  unknown={len(unknown)}")
            form_data[form]=qs
        except Exception as e:
            entry["status"]="ERROR"; entry["errors"].append(str(e)); print(f"    -> ERROR: {e}")
        if log is not None: log.append(entry)
    if not form_data: return None
    df=build_mapping_dataframe(form_data,shared_meta,answer_key)
    save_to_excel(df,out_path)
    return df


def _clean_semester_name(semester: str) -> str:
    """Turn a folder name like 'S16-checked' or 'Spring 2016' into a clean tag
    for filenames, e.g. 'S16' or 'Spring-2016'."""
    s = semester
    # strip common suffixes
    s = re.sub(r'[-_\s]*checked\s*$', '', s, flags=re.I)
    s = re.sub(r'[-_\s]*results?\s*$', '', s, flags=re.I)
    s = s.strip().strip('-_ ')
    s = re.sub(r'\s+', '-', s)
    s = re.sub(r'[^\w\-]', '', s)
    return s or "Semester"


def _descriptive_basename(semester: str, exam_dir: str, meta: dict = None) -> str:
    """Build an explanatory output base name: '<Semester>_<Exam-N>_Result'.

    Prefers a detected human semester (e.g. 'Spring 2016') when available in
    meta, else falls back to the cleaned folder tag (e.g. 'S16'). The exam
    portion keeps its 'Exam-N' / 'Final' style label from the folder.
    """
    sem_tag = ""
    if meta and meta.get("semester"):
        sem_tag = _clean_semester_name(meta["semester"])
    if not sem_tag:
        sem_tag = _clean_semester_name(semester)
    exam_tag = re.sub(r'[^\w\-]+', '-', exam_dir.strip()).strip('-_')
    exam_tag = re.sub(r'_{2,}', '_', exam_tag)
    return f"{sem_tag}_{exam_tag}_Result"


def write_issues_report(report_path: str, semester: str, exam_dir: str,
                        file_entries: list, df=None,
                        available_forms: list = None) -> bool:
    """Write a human-readable issues report for ONE exam set, to be downloaded
    and kept ALONGSIDE the Excel mapping. Lists everything that prevents the
    Excel from being complete/correct: parse errors, empty files, missing
    question numbers, number-sequence jumps, unmarked answers, and — scanned
    from the generated mapping — every MISSING / N/A / Unknown cell.

    Returns True if any issues were written, False if the exam set is fully
    clean (in which case a short 'all clear' report is still written so the
    reviewer has positive confirmation).
    """
    lines = []
    lines.append("=" * 68)
    lines.append(f"  ISSUES REPORT — {semester} / {exam_dir}")
    lines.append("  Keep this file with the Excel mapping. It lists everything")
    lines.append("  that may need a manual check before the mapping is final.")
    lines.append("=" * 68)
    lines.append("")

    any_issue = False

    # ── Omit notes that could not be auto-attached (need a human decision) ─────
    review_omit_entries = [(e.get("form","?"), e.get("file","?"), e.get("review_omits",[]))
                           for e in file_entries if e.get("review_omits")]
    if review_omit_entries:
        any_issue = True
        total = sum(len(n) for _,_,n in review_omit_entries)
        lines.append(f"OMIT NOTES NEEDING MANUAL REVIEW ({total}) — a 'not graded / "
                     "credit to all' note was found but could NOT be confidently")
        lines.append("  tied to a specific question (it floats at the end of the exam")
        lines.append("  or beside a different question). Check the PDF and set OMIT=Yes")
        lines.append("  by hand on the correct master row(s):")
        for form, fname, notes in review_omit_entries:
            for qn, excerpt in notes:
                clean_ex = re.sub(r'\s+', ' ', excerpt).strip()
                lines.append(f"    Form {form} (near Q{qn}): \"{clean_ex}\"")
        lines.append("")


    for e in file_entries:
        flags = []
        if e.get("status") not in ("OK", None):
            flags.append(f"status={e['status']}")
        if e.get("gaps"):
            flags.append(f"missing question numbers {e['gaps']}")
        if e.get("jumps"):
            flags.append("number jumps: " +
                         "; ".join(f"{a} then {b}" for a, b in e["jumps"]))
        if e.get("mark_unknown"):
            flags.append(f"{len(e['mark_unknown'])} MC question(s) with no "
                         f"detectable answer mark: Q{', Q'.join(e['mark_unknown'])}")
        if e.get("img_unknown"):
            flags.append(f"{len(e['img_unknown'])} possible image/diagram "
                         f"question(s) — choices unreadable, needs review: "
                         f"Q{', Q'.join(e['img_unknown'])}")
        if e.get("errors"):
            for er in e["errors"]:
                flags.append(f"ERROR: {er}")
        if flags:
            any_issue = True
            lines.append(f"FORM {e.get('form','?')}  ({e.get('file','?')})")
            lines.append(f"    parsed {e.get('questions',0)} question(s)")
            for fl in flags:
                lines.append(f"    - {fl}")
            # also surface free-response (expected, no action usually needed)
            for w in e.get("warnings", []):
                if "free-response" in w:
                    lines.append(f"    - (expected) {w}")
            lines.append("")

    # ── Mapping-stage issues: explain EVERY incomplete cell in the Excel ──────
    # We cross-reference the per-form parse classification so each Unknown /
    # N/A cell carries a plain-English reason a reviewer can act on:
    #   • image-choice  -> "possible image/diagram — choices not readable"
    #   • free-response -> "free-response question — no A-E choices"
    #   • unmarked      -> "answer key mark not detected"
    #   • MISSING       -> "this form has no counterpart for the question"
    # Build {(form, q_num_str): reason} from the per-file entries.
    reason_by_cell = {}
    for e in file_entries:
        form = e.get("form", "?")
        for qn in e.get("img_unknown", []) or []:
            reason_by_cell[(form, str(qn))] = "possible image/diagram — answer choices could not be read"
        for qn in e.get("fr_unknown", []) or []:
            reason_by_cell[(form, str(qn))] = "free-response question — no multiple-choice options"
        for qn in e.get("mark_unknown", []) or []:
            reason_by_cell[(form, str(qn))] = "answer key mark not detected (key may be unmarked)"

    if df is not None and available_forms:
        incomplete = []   # (mnum, form, kind, detail, prompt)
        for _, row in df.iterrows():
            mnum = row.get("Master Question Number", "?")
            prompt = str(row.get("Question Prompt", ""))[:55]
            variant = str(row.get("Variant", "") or "")
            for form in available_forms:
                num_cell = str(row.get(f"Form {form} #", ""))
                ans_cell = str(row.get(f"Form {form} Correct Choice", ""))
                if num_cell == "MISSING":
                    # A real gap: this question exists on other forms but the
                    # matcher found no counterpart here. Worth a manual look.
                    incomplete.append((mnum, form, "MISSING",
                        "this form has no counterpart for this question", prompt))
                elif num_cell == "N/A":
                    # N/A on a VARIANT row is routine (this form simply has the
                    # other variant) — skip it. N/A on a NON-variant row is
                    # unusual and worth flagging.
                    if not variant:
                        incomplete.append((mnum, form, "N/A",
                            "this form has no entry for this question", prompt))
                elif ans_cell == "Unknown":
                    # The question IS on this form but its answer wasn't
                    # resolved — explain WHY using the parse classification.
                    reason = reason_by_cell.get((form, num_cell),
                        "correct answer could not be determined")
                    incomplete.append((mnum, form, "Unknown", reason, prompt))

        if incomplete:
            any_issue = True
            # Group by kind so the report leads with the most actionable items.
            img_cells = [c for c in incomplete if "image" in c[3]]
            mark_cells = [c for c in incomplete if "mark not detected" in c[3]]
            miss_cells = [c for c in incomplete if c[2] == "MISSING"]
            other_cells = [c for c in incomplete
                           if c not in img_cells and c not in mark_cells and c not in miss_cells]

            def _emit(title, cells):
                if not cells:
                    return
                lines.append(f"{title} ({len(cells)} cell(s)):")
                for mnum, form, kind, detail, prompt in cells:
                    lines.append(f"    Master Q{mnum}  Form {form} = {kind}  —  {detail}")
                    lines.append(f"        Q: {prompt}")
                lines.append("")

            _emit("POSSIBLE IMAGE / DIAGRAM — answer choices could not be read, "
                  "verify manually against the PDF", img_cells)
            _emit("ANSWER MARK NOT DETECTED — the key may be unmarked here, "
                  "confirm the correct choice", mark_cells)
            _emit("MISSING — a form has no counterpart for the question", miss_cells)
            _emit("OTHER UNRESOLVED CELLS", other_cells)

    if not any_issue:
        lines.append("✅ No issues detected. Every form parsed completely, every")
        lines.append("   question number is present and in sequence, and every")
        lines.append("   answer was detected. This mapping looks complete.")
        lines.append("")

    lines.append("=" * 68)
    lines.append("END OF REPORT")

    with open(report_path, "w") as rf:
        rf.write("\n".join(lines))
    return any_issue


def process_all_semesters(base_path,answer_key=None,log_path=None):
    import datetime
    all_log=[]; exam_log=[]
    for semester in sorted(os.listdir(base_path)):
        sem_path=os.path.join(base_path,semester)
        if not os.path.isdir(sem_path): continue
        for exam_dir in sorted(os.listdir(sem_path)):
            exam_path=os.path.join(sem_path,exam_dir)
            if not os.path.isdir(exam_path) or not re.search(r'exam|midterm|final',exam_dir,re.I): continue
            pdf_files=[f for f in os.listdir(exam_path) if f.lower().endswith('.pdf')]
            lm,_=assign_form_letters(pdf_files)
            pm={letter:os.path.join(exam_path,f) for f,letter in lm.items()}
            if len(pm)<1:
                exam_log.append({"semester":semester,"exam":exam_dir,"status":"SKIPPED","forms":0,"note":"No recognisable form PDFs found"})
                continue
            print(f"\nProcessing: {semester} / {exam_dir}")
            rd=os.path.join(sem_path,"Result"); os.makedirs(rd,exist_ok=True)
            base=_descriptive_basename(semester, exam_dir)
            out=os.path.join(rd, f"{base}.xlsx")
            report_path=os.path.join(rd, f"{base}_ISSUES.txt")
            fe=[]
            try:
                df_out=process_pdfs(pm,meta_override={"semester":semester,"exam":exam_dir},
                             answer_key=answer_key,out_path=out,log=fe)
                # Write the downloadable issues report next to the Excel
                avail=sorted(pm.keys())
                try:
                    had_issues=write_issues_report(report_path, semester, exam_dir,
                                                   fe, df=df_out, available_forms=avail)
                    print(f"  Issues report -> {report_path}"
                          + ("" if had_issues else "  (no issues)"))
                except Exception as _re:
                    print(f"  (could not write issues report: {_re})")
                all_log.extend(fe)
                sts=[e["status"] for e in fe]
                es=("✅ PASSED" if all(s=="OK" for s in sts)
                    else "❌ ERROR" if any(s=="ERROR" for s in sts)
                    else "❌ EMPTY" if any(s=="EMPTY" for s in sts)
                    else "⚠️  CHECK")
                exam_log.append({"semester":semester,"exam":exam_dir,"status":es,"forms":len(fe),
                    "total_q":sum(e["questions"] for e in fe),"unknown":sum(e["unknown"] for e in fe),
                    "warnings":sum(len(e["warnings"]) for e in fe),"errors":sum(len(e["errors"]) for e in fe),"output":out})
            except Exception as e:
                exam_log.append({"semester":semester,"exam":exam_dir,"status":"❌ CRASH","forms":len(pm),"note":str(e)})

    if log_path is None: log_path=os.path.join(base_path,"extraction_log.txt")
    ts=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    passed=sum(1 for e in exam_log if "PASSED" in e.get("status",""))
    checked=sum(1 for e in exam_log if "CHECK"  in e.get("status",""))
    errors =sum(1 for e in exam_log if "ERROR"  in e.get("status","") or "CRASH" in e.get("status",""))
    skipped=sum(1 for e in exam_log if "SKIPPED" in e.get("status",""))
    with open(log_path,"w") as lf:
        lf.write("="*70+f"\n  EXAM EXTRACTION LOG\n  Generated: {ts}\n  Base path: {base_path}\n"+"="*70+"\n\n")
        lf.write("SUMMARY\n"+"-"*40+"\n")
        lf.write(f"  Total exam sets processed : {len(exam_log)}\n  ✅ Fully passed            : {passed}\n")
        lf.write(f"  ⚠️  Need checking           : {checked}\n  ❌ Errors / crashes        : {errors}\n")
        lf.write(f"  ⏭️  Skipped (no PDFs)       : {skipped}\n  Total files processed      : {len(all_log)}\n")
        lf.write(f"  Total questions extracted  : {sum(e.get('questions',0) for e in all_log)}\n\n")
        ji=[e for e in all_log if e.get("jumps")]
        lf.write("PER-FORM NUMBER CONTINUITY\n"+"-"*40+"\n")
        if not ji: lf.write("  ✅ All forms have unbroken question numbering (no jumps like 5→7 or 15→17).\n\n")
        else:
            lf.write(f"  ❌ {len(ji)} form(s) with BROKEN sequence:\n")
            for e in ji: lf.write(f"      {e['semester']}/{e['exam']}/Form {e['form']} ({e['file']}): "+"; ".join(f"{a} then {b}" for a,b in e["jumps"])+"\n")
            lf.write("\n")
        gi=[e for e in all_log if e.get("gaps")]; mi=[e for e in all_log if e.get("mark_unknown")]
        if gi or mi:
            lf.write("MANUAL REVIEW WORKLIST\n"+"-"*40+"\n  (Only actionable items — missing questions and unmarked\n   multiple-choice answers.)\n\n")
            if gi:
                lf.write(f"  ▸ Missing question numbers ({sum(len(e['gaps']) for e in gi)} across {len(gi)} file(s)):\n")
                for e in gi: lf.write(f"      {e['semester']}/{e['exam']}/{e['form']}: {e['gaps']}  ({e['file']})\n")
                lf.write("\n")
            if mi:
                lf.write(f"  ▸ MC questions needing a manual answer ({sum(len(e['mark_unknown']) for e in mi)} across {len(mi)} file(s)):\n")
                for e in mi: lf.write(f"      {e['semester']}/{e['exam']}/{e['form']}: Q{', Q'.join(e['mark_unknown'])}\n")
                lf.write("\n\n")
        lf.write("EXAM SET RESULTS\n"+"-"*70+"\n"); cs=None
        for e in exam_log:
            if e["semester"]!=cs: lf.write(f"\n  [{e['semester']}]\n"); cs=e["semester"]
            lf.write(f"    {e['status']:15s}  {e['exam']:20s}")
            q=e.get("total_q",0); u=e.get("unknown",0); w=e.get("warnings",0); er=e.get("errors",0); nt=e.get("note","")
            if q:  lf.write(f"  {q} Qs")
            if u:  lf.write(f"  {u} unknown")
            if w:  lf.write(f"  {w} warnings")
            if er: lf.write(f"  {er} errors")
            if nt: lf.write(f"  — {nt}")
            lf.write("\n")
        issues=[e for e in all_log if e["warnings"] or e["errors"] or e["status"]!="OK"]
        if issues:
            lf.write("\n\nFILES NEEDING ATTENTION\n"+"-"*70+"\n")
            for e in issues:
                lf.write(f"\n  {e['status']:8s}  {e['file']}\n           Semester: {e['semester'] or 'NOT DETECTED'}  Exam: {e['exam'] or 'NOT DETECTED'}  Form: {e['form']}  Questions: {e['questions']}\n")
                for w in e["warnings"]: lf.write(f"           ⚠️  {w}\n")
                for er in e["errors"]:  lf.write(f"           ❌ {er}\n")
        else: lf.write("\n\nAll files passed without issues.\n")
        lf.write("\n"+"="*70+"\nEND OF LOG\n")
    print(f"\n{'='*50}\nLog written → {log_path}\n  ✅ Passed: {passed}  ⚠️ Check: {checked}  ❌ Error: {errors}")
    return exam_log


if __name__ == "__main__":
    ak=ANSWER_KEY_FORM_A or None
    if len(sys.argv)==1:
        process_all_semesters("data",ak,log_path="data/extraction_log.txt")
    elif len(sys.argv)==2 and os.path.isdir(sys.argv[1]):
        process_all_semesters(sys.argv[1],ak,log_path=os.path.join(sys.argv[1],"extraction_log.txt"))
    elif len(sys.argv)>=3:
        paths=sys.argv[1:]
        lm,unr=assign_form_letters([os.path.basename(p) for p in paths])
        if unr:
            print("Cannot determine form letter for: "+", ".join(unr)+"\nName files like ExamA.pdf, FormB.pdf, vEarly A.pdf, etc.")
            sys.exit(1)
        by_base={os.path.basename(p):p for p in paths}
        pm={letter:by_base[fname] for fname,letter in lm.items()}
        fe=[]
        df_out=process_pdfs(pm,{},ak,"Exam_Mapping.xlsx",log=fe)
        # Rename to a descriptive base once metadata is known, and write the
        # downloadable issues report alongside the Excel.
        meta_now={"semester":fe[0]["semester"]} if fe and fe[0].get("semester") else None
        exam_now=(fe[0].get("exam") or "Exam") if fe else "Exam"
        sem_dir=meta_now["semester"] if meta_now else "Result"
        base=_descriptive_basename(sem_dir, exam_now, meta_now)
        out=f"{base}.xlsx"
        try:
            if os.path.exists("Exam_Mapping.xlsx"): os.replace("Exam_Mapping.xlsx",out)
        except Exception:
            out="Exam_Mapping.xlsx"
        report_path=f"{base}_ISSUES.txt"
        try:
            had=write_issues_report(report_path, sem_dir, exam_now, fe,
                                    df=df_out, available_forms=sorted(pm.keys()))
            print(f"Saved -> {out}\nIssues report -> {report_path}"
                  + ("" if had else "  (no issues)"))
        except Exception as _re:
            print(f"Saved -> {out}  (could not write issues report: {_re})")
    else:
        print(__doc__)